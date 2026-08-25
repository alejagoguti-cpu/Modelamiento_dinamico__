#!/usr/bin/env python3
"""
Exportar datos del POT a Excel
"""

import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

# Cargar datos
with open('data/pot_nodos_completos.json', 'r', encoding='utf-8') as f:
    pot_data = json.load(f)

# Crear workbook
wb = Workbook()
wb.remove(wb.active)  # Remover hoja por defecto

# Estilos
header_fill = PatternFill(start_color="2FBFAE", end_color="2FBFAE", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
alt_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def style_header(ws):
    """Aplicar estilo a encabezados"""
    for cell in ws[1]:
        if cell.value:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

def style_cells(ws):
    """Aplicar estilo a celdas"""
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if idx % 2 == 0:
                cell.fill = alt_fill

# ===== HOJA 1: RESUMEN =====
ws_summary = wb.create_sheet("📊 Resumen", 0)
ws_summary['A1'] = "RAPOT - Modelamiento Dinámico del POT Bogotá"
ws_summary['A1'].font = Font(bold=True, size=14, color="2FBFAE")
ws_summary['A3'] = "Generado:"
ws_summary['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

row = 5
for structure_name, structure_data in pot_data['estructuras_territoriales'].items():
    ws_summary[f'A{row}'] = structure_data['nombre']
    ws_summary[f'A{row}'].font = Font(bold=True, color="2FBFAE")
    ws_summary[f'A{row+1}'] = "Componentes:"

    comp_row = row + 2
    for comp_name, comp_data in structure_data.get('componentes', {}).items():
        elementos = comp_data.get('elementos', [])
        elem_count = len(elementos) if isinstance(elementos, list) else len(elementos.get('elementos', []))

        ws_summary[f'A{comp_row}'] = f"  • {comp_data['nombre']}"
        ws_summary[f'B{comp_row}'] = f"{elem_count} elementos"
        comp_row += 1

    row = comp_row + 2

ws_summary.column_dimensions['A'].width = 50
ws_summary.column_dimensions['B'].width = 20

# ===== HOJA 2: ESTRUCTURAS =====
ws_struct = wb.create_sheet("Estructuras", 1)
ws_struct.append(['ID', 'Nombre', 'Color', 'Descripción'])
style_header(ws_struct)

for key, struct in pot_data['estructuras_territoriales'].items():
    ws_struct.append([
        key.upper(),
        struct.get('nombre', ''),
        struct.get('color', ''),
        struct.get('descripcion', '')
    ])

for col in ws_struct.columns:
    ws_struct.column_dimensions[col[0].column_letter].width = 25
style_cells(ws_struct)

# ===== HOJAS POR ESTRUCTURA =====
for est_key, estructura in pot_data['estructuras_territoriales'].items():
    est_name = estructura['nombre'].split()[0][:10]

    # Hoja de componentes
    ws_comp = wb.create_sheet(f"{est_name} - Componentes", len(wb.sheetnames))
    ws_comp.append(['Componente', 'Descripción', 'Cantidad de Elementos'])
    style_header(ws_comp)

    for comp_key, comp in estructura.get('componentes', {}).items():
        elementos = comp.get('elementos', [])
        elem_count = len(elementos) if isinstance(elementos, list) else len(elementos.get('elementos', []))
        ws_comp.append([
            comp.get('nombre', ''),
            comp.get('descripcion', ''),
            elem_count
        ])

    ws_comp.column_dimensions['A'].width = 35
    ws_comp.column_dimensions['B'].width = 40
    ws_comp.column_dimensions['C'].width = 20
    style_cells(ws_comp)

    # Hoja de elementos
    ws_elem = wb.create_sheet(f"{est_name} - Elementos", len(wb.sheetnames))
    ws_elem.append(['Componente', 'Elemento', 'Tipo'])
    style_header(ws_elem)

    for comp_key, comp in estructura.get('componentes', {}).items():
        elementos = comp.get('elementos', [])
        elementos_list = elementos if isinstance(elementos, list) else elementos.get('elementos', [])

        for elem in elementos_list:
            elem_name = elem.get('nombre') if isinstance(elem, dict) else str(elem)
            ws_elem.append([
                comp.get('nombre', ''),
                elem_name,
                comp_key
            ])

    ws_elem.column_dimensions['A'].width = 35
    ws_elem.column_dimensions['B'].width = 45
    ws_elem.column_dimensions['C'].width = 20
    style_cells(ws_elem)

# ===== HOJA: TODAS LAS RELACIONES =====
ws_relations = wb.create_sheet("Relaciones", 2)
ws_relations.append(['De', 'A', 'Relación', 'Cita', 'Página'])
style_header(ws_relations)

mapping = {
    'e1': ['EEP', 'ESECI'],
    'e2': ['EIP', 'EEP'],
    'e3': ['EFC', 'EEP'],
    'e4': ['EFC', 'ESECI'],
    'e5': ['EIP', 'EFC'],
    'e6': ['EIP', 'ESECI']
}

for rel_key, rel_data in pot_data.get('relations', {}).items():
    if rel_key in mapping:
        from_est, to_est = mapping[rel_key]
        ws_relations.append([
            from_est,
            to_est,
            rel_data.get('label', ''),
            rel_data.get('quote', ''),
            rel_data.get('page', '')
        ])

ws_relations.column_dimensions['A'].width = 15
ws_relations.column_dimensions['B'].width = 15
ws_relations.column_dimensions['C'].width = 30
ws_relations.column_dimensions['D'].width = 60
ws_relations.column_dimensions['E'].width = 12
style_cells(ws_relations)

# Guardar
output_file = f'RAPOT_Datos_POT_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
wb.save(output_file)

print(f'✅ Excel creado: {output_file}')
print(f'📊 Hojas: {len(wb.sheetnames)}')
print(f'📝 Hojas:')
for sheet in wb.sheetnames:
    print(f'   • {sheet}')
